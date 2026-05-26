"""XLSX parser.

For each sheet, emits one chunk per contiguous data region (a "region" is
the bounding box of non-empty cells). Text representation: tab-separated
rows, newline-separated rows. Locator carries the sheet name and A1-style
cell range so citations can point at the exact region.

For pharma data (AE listings, demographic tables, lab summaries) this
parser is intentionally simple — the production pipeline relies on
deterministic named queries against the source warehouse, not LLM
interpretation of XLSX dumps. This exists for the cases where a study
report ships its data as an .xlsx attachment.
"""

from __future__ import annotations

import io
from collections.abc import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from shared.schemas import CanonicalDocument, ChunkKind, ParsedChunk, XlsxLocator

from .parser import hash_text, new_chunk_id


class XlsxParser:
    mime_types = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    )

    def parse(self, doc: CanonicalDocument, raw: bytes) -> Iterable[ParsedChunk]:
        workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        for sheet in workbook.worksheets:
            region = _bounding_region(sheet)
            if region is None:
                continue
            min_row, min_col, max_row, max_col = region
            text = _region_to_text(sheet, min_row, min_col, max_row, max_col)
            if not text.strip():
                continue
            cell_range = (
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            )
            yield ParsedChunk(
                chunk_id=new_chunk_id(),
                source_doc_id=doc.doc_id,
                source_doc_version=doc.doc_version,
                kind=ChunkKind.SHEET_REGION,
                text=text,
                text_hash=hash_text(text),
                char_count=len(text),
                locator=XlsxLocator(sheet=sheet.title, cell_range=cell_range),
                tags=list(doc.tags),
            )


def _bounding_region(sheet: Worksheet) -> tuple[int, int, int, int] | None:
    """Return (min_row, min_col, max_row, max_col) of the populated area,
    or None if the sheet is empty. Uses 1-based indices.
    """
    min_row = sheet.min_row
    min_col = sheet.min_column
    max_row = sheet.max_row
    max_col = sheet.max_column
    if not (min_row and min_col and max_row and max_col):
        return None
    if max_row < min_row or max_col < min_col:
        return None
    # openpyxl reports min/max optimistically; verify at least one non-empty cell exists.
    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
    ):
        if any(cell not in (None, "") for cell in row):
            return min_row, min_col, max_row, max_col
    return None


def _region_to_text(
    sheet: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int
) -> str:
    """Render a sheet region as tab-separated rows joined with newlines."""
    lines: list[str] = []
    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
    ):
        cells = ["" if v is None else str(v) for v in row]
        if any(c.strip() for c in cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)
